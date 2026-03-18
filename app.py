from flask import Flask, render_template, request, jsonify, redirect, url_for, session, send_file, make_response
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from flask_socketio import SocketIO, emit, join_room, leave_room
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from datetime import datetime, timedelta
import json
import os
import re
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from io import BytesIO
import tempfile

# Always load templates from this project folder (avoids stale/wrong UI when cwd differs)
_APP_DIR = os.path.dirname(os.path.abspath(__file__))
app = Flask(__name__, template_folder=os.path.join(_APP_DIR, 'templates'))
app.config['SECRET_KEY'] = 'your-secret-key-change-in-production'
app.config['TEMPLATES_AUTO_RELOAD'] = True
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///classroom_app.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
socketio = SocketIO(app, cors_allowed_origins="*", async_mode="threading")

# Database Models
class Professor(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class Student(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    student_number = db.Column(db.String(50), unique=True, nullable=False)
    first_name = db.Column(db.String(100), nullable=False)
    preferred_name = db.Column(db.String(100), nullable=True)
    last_name = db.Column(db.String(100), nullable=False)
    email = db.Column(db.String(120), nullable=False)
    rfid_card_id = db.Column(db.String(100), unique=True, nullable=True)
    password_hash = db.Column(db.String(255), nullable=True)  # Nullable for first-time setup
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class Class(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    professor_id = db.Column(db.Integer, db.ForeignKey('professor.id'), nullable=False)
    name = db.Column(db.String(200), nullable=False)
    class_code = db.Column(db.String(20), unique=True, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    is_active = db.Column(db.Boolean, default=False)
    professor = db.relationship('Professor', backref=db.backref('classes', lazy=True))

class Enrollment(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    class_id = db.Column(db.Integer, db.ForeignKey('class.id'), nullable=False)
    student_id = db.Column(db.Integer, db.ForeignKey('student.id'), nullable=False)
    enrolled_at = db.Column(db.DateTime, default=datetime.utcnow)
    is_active = db.Column(db.Boolean, default=True, nullable=False)
    class_obj = db.relationship('Class', backref=db.backref('enrollments', lazy=True))
    student = db.relationship('Student', backref=db.backref('enrollments', lazy=True))

class Attendance(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    class_id = db.Column(db.Integer, db.ForeignKey('class.id'), nullable=False)
    student_id = db.Column(db.Integer, db.ForeignKey('student.id'), nullable=False)
    date = db.Column(db.Date, nullable=False)
    present = db.Column(db.Boolean, default=True)
    join_time = db.Column(db.DateTime, nullable=True)  # Time when student joined
    leave_time = db.Column(db.DateTime, nullable=True)  # Time when student left (early logout or class end)
    timestamp = db.Column(db.DateTime, default=datetime.utcnow)
    class_obj = db.relationship('Class', backref=db.backref('attendances', lazy=True))
    student = db.relationship('Student', backref=db.backref('attendances', lazy=True))

class Participation(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    class_id = db.Column(db.Integer, db.ForeignKey('class.id'), nullable=False)
    student_id = db.Column(db.Integer, db.ForeignKey('student.id'), nullable=False)
    date = db.Column(db.Date, nullable=False)
    peer_grade = db.Column(db.Float, default=0.0)
    instructor_grade = db.Column(db.Float, default=0.0)
    hand_raises = db.Column(db.Integer, default=0)
    thumbs_up = db.Column(db.Integer, default=0)
    thumbs_down = db.Column(db.Integer, default=0)
    class_obj = db.relationship('Class', backref=db.backref('participations', lazy=True))
    student = db.relationship('Student', backref=db.backref('participations', lazy=True))

class Poll(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    class_id = db.Column(db.Integer, db.ForeignKey('class.id'), nullable=False)
    question = db.Column(db.String(500), nullable=False)
    options = db.Column(db.Text, nullable=False)  # JSON string
    correct_answer = db.Column(db.Integer, nullable=True)
    is_graded = db.Column(db.Boolean, default=False)  # Whether this poll counts toward grade
    is_anonymous = db.Column(db.Boolean, default=False)
    is_active = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    class_obj = db.relationship('Class', backref=db.backref('polls', lazy=True))

class PollResponse(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    poll_id = db.Column(db.Integer, db.ForeignKey('poll.id'), nullable=False)
    student_id = db.Column(db.Integer, db.ForeignKey('student.id'), nullable=False)
    answer = db.Column(db.Integer, nullable=False)
    is_correct = db.Column(db.Boolean, default=False)
    timestamp = db.Column(db.DateTime, default=datetime.utcnow)
    poll = db.relationship('Poll', backref=db.backref('responses', lazy=True))
    student = db.relationship('Student', backref=db.backref('poll_responses', lazy=True))

class ClassSettings(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    class_id = db.Column(db.Integer, db.ForeignKey('class.id'), nullable=False, unique=True)
    show_first_name_only = db.Column(db.Boolean, default=False)
    quiet_mode = db.Column(db.Boolean, default=False)
    class_obj = db.relationship('Class', backref=db.backref('settings', uselist=False))

class ClassSession(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    class_id = db.Column(db.Integer, db.ForeignKey('class.id'), nullable=False)
    start_time = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)
    end_time = db.Column(db.DateTime, nullable=True)
    exclude_from_grading = db.Column(db.Boolean, default=False)  # If True, this session doesn't count toward attendance grades
    class_obj = db.relationship('Class', backref=db.backref('sessions', lazy=True))

class GradingWeights(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    class_id = db.Column(db.Integer, db.ForeignKey('class.id'), nullable=False, unique=True)
    attendance_weight = db.Column(db.Float, default=25.0, nullable=False)
    instructor_participation_weight = db.Column(db.Float, default=25.0, nullable=False)
    peer_participation_weight = db.Column(db.Float, default=25.0, nullable=False)
    poll_weight = db.Column(db.Float, default=25.0, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    class_obj = db.relationship('Class', backref=db.backref('grading_weights', uselist=False))

class HandRaise(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    class_id = db.Column(db.Integer, db.ForeignKey('class.id'), nullable=False)
    student_id = db.Column(db.Integer, db.ForeignKey('student.id'), nullable=False)
    timestamp = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    cleared = db.Column(db.Boolean, default=False, nullable=False)
    class_obj = db.relationship('Class', backref=db.backref('hand_raises', lazy=True))
    student = db.relationship('Student', backref=db.backref('hand_raises', lazy=True))

@login_manager.user_loader
def load_user(user_id):
    return Professor.query.get(int(user_id))

# Routes
@app.route('/')
def index():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        user_type = request.form.get('user_type', 'professor')
        
        if user_type == 'professor':
            professor = Professor.query.filter_by(username=username).first()
            if professor and check_password_hash(professor.password_hash, password):
                login_user(professor)
                return jsonify({'success': True, 'redirect': url_for('dashboard')})
            return jsonify({'success': False, 'error': 'Invalid credentials'})
        else:
            # Student login will be handled differently
            return jsonify({'success': False, 'error': 'Use student interface'})
    
    return render_template('login.html')

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        data = request.get_json() if request.is_json else request.form.to_dict()
        username = data.get('username')
        email = data.get('email')
        password = data.get('password')
        
        if not username or not email or not password:
            return jsonify({'success': False, 'error': 'All fields are required'})
        
        # Check if username already exists
        if Professor.query.filter_by(username=username).first():
            return jsonify({'success': False, 'error': 'Username already exists'})
        
        # Check if email already exists
        if Professor.query.filter_by(email=email).first():
            return jsonify({'success': False, 'error': 'Email already exists'})
        
        # Create new professor
        professor = Professor(
            username=username,
            email=email,
            password_hash=generate_password_hash(password)
        )
        db.session.add(professor)
        db.session.commit()
        
        # Auto-login the new professor
        login_user(professor)
        
        return jsonify({'success': True, 'redirect': url_for('dashboard')})
    
    return render_template('register.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

@app.route('/dashboard')
@login_required
def dashboard():
    classes = Class.query.filter_by(professor_id=current_user.id).all()
    return render_template('dashboard.html', classes=classes)

@app.route('/preferences')
@login_required
def preferences():
    return render_template('preferences.html')

@app.route('/classroom/<int:class_id>')
@login_required
def classroom(class_id):
    class_obj = Class.query.get_or_404(class_id)
    if class_obj.professor_id != current_user.id:
        return redirect(url_for('dashboard'))
    
    students = db.session.query(Student).join(Enrollment).filter(
        Enrollment.class_id == class_id
    ).all()
    
    settings = ClassSettings.query.filter_by(class_id=class_id).first()
    if not settings:
        settings = ClassSettings(class_id=class_id)
        db.session.add(settings)
        db.session.commit()
    
    # Get active session for attendance grading setting
    active_session = ClassSession.query.filter_by(
        class_id=class_id,
        end_time=None
    ).order_by(ClassSession.start_time.desc()).first()
    
    return render_template('classroom.html', class_obj=class_obj, students=students, settings=settings, active_session=active_session)

@app.route('/classroom/<int:class_id>/students')
@login_required
def students_list(class_id):
    class_obj = Class.query.get_or_404(class_id)
    if class_obj.professor_id != current_user.id:
        return redirect(url_for('dashboard'))
    
    # Get active students
    active_students = db.session.query(Student).join(Enrollment).filter(
        Enrollment.class_id == class_id,
        Enrollment.is_active == True
    ).order_by(Student.last_name, Student.first_name).all()
    
    # Get inactive students
    inactive_students = db.session.query(Student).join(Enrollment).filter(
        Enrollment.class_id == class_id,
        Enrollment.is_active == False
    ).order_by(Student.last_name, Student.first_name).all()
    
    return render_template('students_list.html', class_obj=class_obj, active_students=active_students, inactive_students=inactive_students)

@app.route('/classroom/<int:class_id>/class_data')
@login_required
def class_data(class_id):
    class_obj = Class.query.get_or_404(class_id)
    if class_obj.professor_id != current_user.id:
        return redirect(url_for('dashboard'))
    
    return render_template('class_data.html', class_obj=class_obj)

@app.route('/api/start_class/<int:class_id>', methods=['POST'])
@login_required
def start_class(class_id):
    class_obj = Class.query.get_or_404(class_id)
    if class_obj.professor_id != current_user.id:
        return jsonify({'success': False, 'error': 'Unauthorized'})
    
    class_obj.is_active = True
    
    # Get exclude_from_grading from request if provided
    data = request.get_json() or {}
    exclude_from_grading = data.get('exclude_from_grading', False)
    
    # Create a new class session
    session_record = ClassSession(
        class_id=class_id,
        start_time=datetime.utcnow(),
        exclude_from_grading=exclude_from_grading
    )
    db.session.add(session_record)
    db.session.commit()
    
    socketio.emit('class_started', {'class_id': class_id, 'class_code': class_obj.class_code}, room=f'class_{class_id}')
    
    return jsonify({'success': True, 'redirect': url_for('faculty_dashboard', class_id=class_id)})

@app.route('/api/stop_class/<int:class_id>', methods=['POST'])
@login_required
def stop_class(class_id):
    class_obj = Class.query.get_or_404(class_id)
    if class_obj.professor_id != current_user.id:
        return jsonify({'success': False, 'error': 'Unauthorized'})
    
    class_obj.is_active = False
    
    # Update the active session with end time
    active_session = ClassSession.query.filter_by(
        class_id=class_id,
        end_time=None
    ).order_by(ClassSession.start_time.desc()).first()
    
    end_time = datetime.utcnow()
    if active_session:
        active_session.end_time = end_time
    
    # Log leave_time for all students who are present (auto-logout)
    today = datetime.utcnow().date()
    attendances = Attendance.query.filter_by(
        class_id=class_id,
        date=today,
        present=True
    ).all()
    
    for attendance in attendances:
        # Only set leave_time if not already set (in case of early logout)
        if not attendance.leave_time:
            attendance.leave_time = end_time
    
    db.session.commit()
    
    # Update gradebook with participation data
    update_gradebook(class_id)
    
    socketio.emit('class_stopped', {'class_id': class_id}, room=f'class_{class_id}')
    
    return jsonify({'success': True})

def update_gradebook(class_id):
    today = datetime.utcnow().date()
    students = db.session.query(Student).join(Enrollment).filter(
        Enrollment.class_id == class_id
    ).all()
    
    for student in students:
        participation = Participation.query.filter_by(
            class_id=class_id,
            student_id=student.id,
            date=today
        ).first()
        
        if not participation:
            participation = Participation(
                class_id=class_id,
                student_id=student.id,
                date=today
            )
            db.session.add(participation)
        
        # Graded polls only (matches gradebook)
        poll_responses = PollResponse.query.join(Poll).filter(
            Poll.class_id == class_id,
            PollResponse.student_id == student.id,
            Poll.created_at >= datetime.combine(today, datetime.min.time()),
            Poll.is_graded == True,
        ).all()
        
        poll_grade = 0.0
        if poll_responses:
            correct_count = sum(1 for pr in poll_responses if pr.is_correct)
            poll_grade = (correct_count / len(poll_responses)) * 100
        
        db.session.commit()

@app.route('/faculty_dashboard/<int:class_id>')
@login_required
def faculty_dashboard(class_id):
    class_obj = Class.query.get_or_404(class_id)
    if class_obj.professor_id != current_user.id:
        return redirect(url_for('dashboard'))
    
    students = db.session.query(Student).join(Enrollment).filter(
        Enrollment.class_id == class_id
    ).all()
    
    active_poll = Poll.query.filter_by(class_id=class_id, is_active=True).first()
    
    return render_template('faculty_dashboard.html', class_obj=class_obj, students=students, active_poll=active_poll)

@app.route('/api/gradebook/<int:class_id>')
@login_required
def get_gradebook(class_id):
    class_obj = Class.query.get_or_404(class_id)
    if class_obj.professor_id != current_user.id:
        return jsonify({'error': 'Unauthorized'}), 403
    
    # Get or create grading weights
    grading_weights = GradingWeights.query.filter_by(class_id=class_id).first()
    if not grading_weights:
        # Create default weights (25% each)
        grading_weights = GradingWeights(
            class_id=class_id,
            attendance_weight=25.0,
            instructor_participation_weight=25.0,
            peer_participation_weight=25.0,
            poll_weight=25.0
        )
        db.session.add(grading_weights)
        db.session.commit()
    
    students = db.session.query(Student).join(Enrollment).filter(
        Enrollment.class_id == class_id,
        Enrollment.is_active == True
    ).all()
    
    gradebook_data = []
    for student in students:
        attendances = Attendance.query.filter_by(
            class_id=class_id,
            student_id=student.id
        ).all()
        
        participations = Participation.query.filter_by(
            class_id=class_id,
            student_id=student.id
        ).all()
        
        # Only count graded polls in poll grade calculation
        poll_responses = PollResponse.query.join(Poll).filter(
            Poll.class_id == class_id,
            PollResponse.student_id == student.id,
            Poll.is_graded == True
        ).all()
        
        # Calculate attendance, excluding sessions marked as exclude_from_grading
        # Get all session dates that count toward grading
        sessions = ClassSession.query.filter_by(class_id=class_id).all()
        graded_session_dates = {s.start_time.date() for s in sessions if not s.exclude_from_grading}
        
        # Count attendance only for sessions that count toward grading
        attendance_count = sum(1 for a in attendances if a.present and a.date in graded_session_dates)
        total_graded_classes = len(graded_session_dates)
        attendance_grade = (attendance_count / total_graded_classes * 100) if total_graded_classes > 0 else 0
        
        avg_peer_grade = sum(p.peer_grade for p in participations) / len(participations) if participations else 0
        avg_instructor_grade = sum(p.instructor_grade for p in participations) / len(participations) if participations else 0
        
        poll_grade = 0
        if poll_responses:
            correct_count = sum(1 for pr in poll_responses if pr.is_correct)
            poll_grade = (correct_count / len(poll_responses)) * 100
        
        # Calculate overall grade using weighted average
        overall_grade = (
            (attendance_grade * grading_weights.attendance_weight / 100) +
            (avg_instructor_grade * grading_weights.instructor_participation_weight / 100) +
            (avg_peer_grade * grading_weights.peer_participation_weight / 100) +
            (poll_grade * grading_weights.poll_weight / 100)
        )
        
        gradebook_data.append({
            'student_id': student.id,
            'student_number': student.student_number,
            'name': f"{student.first_name} {student.last_name}",
            'attendance_grade': round(attendance_grade, 2),
            'peer_participation': round(avg_peer_grade, 2),
            'instructor_participation': round(avg_instructor_grade, 2),
            'poll_grade': round(poll_grade, 2),
            'overall_grade': round(overall_grade, 2)
        })
    
    return jsonify(gradebook_data)

@app.route('/api/export_gradebook/<int:class_id>')
@login_required
def export_gradebook(class_id):
    class_obj = Class.query.get_or_404(class_id)
    if class_obj.professor_id != current_user.id:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    # Get grading weights
    grading_weights = GradingWeights.query.filter_by(class_id=class_id).first()
    if not grading_weights:
        grading_weights = GradingWeights(
            class_id=class_id,
            attendance_weight=25.0,
            instructor_participation_weight=25.0,
            peer_participation_weight=25.0,
            poll_weight=25.0
        )
    
    # Get all students (same logic as get_gradebook)
    students = db.session.query(Student).join(Enrollment).filter(
        Enrollment.class_id == class_id,
        Enrollment.is_active == True
    ).all()
    
    # Create a new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Gradebook"
    
    # Define the header row
    headers = ['Student Name', 'Student Number', 'Email', 'Attendance Grade (%)', 
               'Instructor Participation', 'Peer Participation', 'Poll Grade (%)', 'Overall Grade (%)']
    ws.append(headers)
    
    # Style the header row
    header_fill = PatternFill(start_color="2A1A40", end_color="2A1A40", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    # Calculate grades for each student (same logic as get_gradebook)
    sessions = ClassSession.query.filter_by(class_id=class_id).all()
    graded_session_dates = {s.start_time.date() for s in sessions if not s.exclude_from_grading}
    
    for student in students:
        attendances = Attendance.query.filter_by(
            class_id=class_id,
            student_id=student.id
        ).all()
        
        participations = Participation.query.filter_by(
            class_id=class_id,
            student_id=student.id
        ).all()
        
        poll_responses = PollResponse.query.join(Poll).filter(
            Poll.class_id == class_id,
            PollResponse.student_id == student.id,
            Poll.is_graded == True
        ).all()
        
        attendance_count = sum(1 for a in attendances if a.present and a.date in graded_session_dates)
        total_graded_classes = len(graded_session_dates)
        attendance_grade = (attendance_count / total_graded_classes * 100) if total_graded_classes > 0 else 0
        
        avg_peer_grade = sum(p.peer_grade for p in participations) / len(participations) if participations else 0
        avg_instructor_grade = sum(p.instructor_grade for p in participations) / len(participations) if participations else 0
        
        poll_grade = 0
        if poll_responses:
            correct_count = sum(1 for pr in poll_responses if pr.is_correct)
            poll_grade = (correct_count / len(poll_responses)) * 100
        
        overall_grade = (
            (attendance_grade * grading_weights.attendance_weight / 100) +
            (avg_instructor_grade * grading_weights.instructor_participation_weight / 100) +
            (avg_peer_grade * grading_weights.peer_participation_weight / 100) +
            (poll_grade * grading_weights.poll_weight / 100)
        )
        
        row = [
            f"{student.first_name} {student.last_name}",
            student.student_number,
            student.email if hasattr(student, 'email') else '',
            round(attendance_grade, 2),
            round(avg_instructor_grade, 2),
            round(avg_peer_grade, 2),
            round(poll_grade, 2),
            round(overall_grade, 2)
        ]
        ws.append(row)
    
    # Auto-fit column widths
    from openpyxl.utils import get_column_letter
    for col_idx, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_idx)
        max_length = len(str(header))
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_length * 1.15 + 2, 12), 50)
    
    # Create BytesIO object to store the Excel file in memory
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return send_file(
        excel_buffer,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'gradebook_{class_obj.name.replace(" ", "_")}_{datetime.now().strftime("%Y%m%d")}.xlsx'
    )

@app.route('/api/class_metrics/<int:class_id>')
@login_required
def get_class_metrics(class_id):
    class_obj = Class.query.get_or_404(class_id)
    if class_obj.professor_id != current_user.id:
        return jsonify({'error': 'Unauthorized'}), 403
    
    sessions = ClassSession.query.filter_by(class_id=class_id).order_by(ClassSession.start_time.desc()).all()
    
    # Get all enrolled students for this class
    enrolled_students = db.session.query(Student).join(Enrollment).filter(
        Enrollment.class_id == class_id,
        Enrollment.is_active == True
    ).all()
    enrolled_student_ids = {s.id for s in enrolled_students}
    
    sessions_data = []
    session_number = len(sessions)  # Start from highest number (most recent first)
    
    for session in sessions:
        session_date = session.start_time.date()
        session_start = session.start_time
        session_end = session.end_time if session.end_time else datetime.utcnow()
        
        # Get present students count
        present_students = Attendance.query.filter_by(
            class_id=class_id,
            date=session_date,
            present=True
        ).all()
        attendance_count = len(present_students)
        
        # Calculate attendance percentage based on total enrolled students
        total_enrolled = len(enrolled_students)
        attendance_percentage = (attendance_count / total_enrolled * 100) if total_enrolled > 0 else 0
        
        # Get unique hands raised (count distinct students who raised hands during this session)
        # Count hand raises that occurred during this session
        hand_raises_during_session = HandRaise.query.filter(
            HandRaise.class_id == class_id,
            HandRaise.timestamp >= session_start,
            HandRaise.timestamp <= session_end
        ).all()
        unique_hands_raised = len(set(hr.student_id for hr in hand_raises_during_session))
        
        # Graded polls only — ungraded polls excluded from class metrics
        polls = Poll.query.filter(
            Poll.class_id == class_id,
            Poll.created_at >= session_start,
            Poll.created_at <= session_end,
            Poll.is_graded == True,
        ).all()
        
        poll_results = []
        # Calculate poll vote percentage: unique students who voted / attendance count
        unique_poll_voters = set()
        for poll in polls:
            responses = PollResponse.query.filter_by(poll_id=poll.id).all()
            # Track unique students who voted
            for response in responses:
                unique_poll_voters.add(response.student_id)
            
            option_counts = {}
            options = json.loads(poll.options)
            for i in range(len(options)):
                option_counts[i] = sum(1 for r in responses if r.answer == i)
            
            poll_results.append({
                'poll_id': poll.id,
                'question': poll.question,
                'options': options,
                'option_counts': option_counts,
                'total_responses': len(responses),
                'is_graded': poll.is_graded
            })
        
        # Calculate overall poll vote percentage for the session
        poll_vote_percentage = (len(unique_poll_voters) / attendance_count * 100) if attendance_count > 0 else 0
        
        # Get attendance records for this session date
        attendances = Attendance.query.filter_by(
            class_id=class_id,
            date=session_date
        ).all()
        
        # Create a map of student_id to attendance record
        attendance_map = {att.student_id: att for att in attendances}
        
        # Build attendance list with ALL enrolled students
        attendance_list = []
        for student in enrolled_students:
            att = attendance_map.get(student.id)
            
            if att:
                sign_in_time = att.join_time if att.join_time else att.timestamp
                # Only show sign out time if student left before professor ended the class
                # If class hasn't ended (session_end is None or is current time), show leave_time if it exists
                # If class has ended, only show leave_time if it's before the session end time
                sign_out_time = None
                if att.leave_time:
                    if session.end_time is None:
                        # Class hasn't ended yet, show leave_time if it exists (early logout)
                        sign_out_time = att.leave_time
                    elif att.leave_time < session.end_time:
                        # Class has ended, only show if student left before class ended
                        sign_out_time = att.leave_time
            else:
                sign_in_time = None
                sign_out_time = None
            
            attendance_list.append({
                'student_id': student.id,
                'student_number': student.student_number,
                'student_name': f"{student.preferred_name or student.first_name} {student.last_name}",
                'sign_in_time': sign_in_time.isoformat() if sign_in_time else None,
                'sign_out_time': sign_out_time.isoformat() if sign_out_time else None,
                'present': att.present if att else False
            })
        
        sessions_data.append({
            'session_id': session.id,
            'session_number': session_number,
            'start_time': session.start_time.isoformat(),
            'end_time': session.end_time.isoformat() if session.end_time else None,
            'exclude_from_grading': session.exclude_from_grading,
            'engagement_metrics': {
                'attendance': round(attendance_percentage, 1),
                'unique_hands_raised': unique_hands_raised,
                'poll_vote_percentage': round(poll_vote_percentage, 1)
            },
            'poll_results': poll_results,
            'attendance_list': attendance_list
        })
        
        session_number -= 1  # Decrement for next session
    
    return jsonify(sessions_data)

@app.route('/api/update_poll_grading/<int:poll_id>', methods=['POST'])
@login_required
def update_poll_grading(poll_id):
    poll = Poll.query.get_or_404(poll_id)
    class_obj = Class.query.get_or_404(poll.class_id)
    
    if class_obj.professor_id != current_user.id:
        return jsonify({'error': 'Unauthorized'}), 403
    
    data = request.get_json()
    is_graded = data.get('is_graded', False)
    
    poll.is_graded = is_graded
    db.session.commit()
    
    return jsonify({'success': True, 'is_graded': poll.is_graded})

@app.route('/api/update_session_grading/<int:session_id>', methods=['POST'])
@login_required
def update_session_grading(session_id):
    session = ClassSession.query.get_or_404(session_id)
    class_obj = Class.query.get_or_404(session.class_id)
    
    if class_obj.professor_id != current_user.id:
        return jsonify({'error': 'Unauthorized'}), 403
    
    data = request.get_json()
    exclude_from_grading = data.get('exclude_from_grading', False)
    
    session.exclude_from_grading = exclude_from_grading
    db.session.commit()
    
    return jsonify({'success': True, 'exclude_from_grading': session.exclude_from_grading})

@app.route('/api/delete_session/<int:session_id>', methods=['DELETE'])
@login_required
def delete_session(session_id):
    session = ClassSession.query.get_or_404(session_id)
    class_obj = Class.query.get_or_404(session.class_id)
    
    if class_obj.professor_id != current_user.id:
        return jsonify({'error': 'Unauthorized'}), 403
    
    # Delete the session
    db.session.delete(session)
    db.session.commit()
    
    return jsonify({'success': True})

@app.route('/api/update_settings/<int:class_id>', methods=['POST'])
@login_required
def update_settings(class_id):
    class_obj = Class.query.get_or_404(class_id)
    if class_obj.professor_id != current_user.id:
        return jsonify({'success': False, 'error': 'Unauthorized'})
    
    settings = ClassSettings.query.filter_by(class_id=class_id).first()
    if not settings:
        settings = ClassSettings(class_id=class_id)
        db.session.add(settings)
    
    data = request.get_json()
    
    # Update class name and code if provided
    if 'class_name' in data:
        class_obj.name = data.get('class_name')
    if 'class_code' in data:
        new_class_code = data.get('class_code')
        # Check if class code is unique (excluding current class)
        existing_class = Class.query.filter_by(class_code=new_class_code).first()
        if existing_class and existing_class.id != class_id:
            return jsonify({'success': False, 'error': 'Class code already exists'})
        class_obj.class_code = new_class_code
    
    # Update settings
    settings.show_first_name_only = data.get('show_first_name_only', False)
    settings.quiet_mode = data.get('quiet_mode', False)
    
    # Update current session's exclude_from_grading if provided
    if 'exclude_from_grading' in data:
        active_session = ClassSession.query.filter_by(
            class_id=class_id,
            end_time=None
        ).order_by(ClassSession.start_time.desc()).first()
        
        if active_session:
            active_session.exclude_from_grading = data.get('exclude_from_grading', False)
    
    db.session.commit()
    
    socketio.emit('settings_updated', {
        'show_first_name_only': settings.show_first_name_only,
        'quiet_mode': settings.quiet_mode,
        'exclude_from_grading': active_session.exclude_from_grading if active_session else False
    }, room=f'class_{class_id}')
    
    return jsonify({'success': True})

@app.route('/api/grading_weights/<int:class_id>', methods=['GET'])
@login_required
def get_grading_weights(class_id):
    class_obj = Class.query.get_or_404(class_id)
    if class_obj.professor_id != current_user.id:
        return jsonify({'error': 'Unauthorized'}), 403
    
    grading_weights = GradingWeights.query.filter_by(class_id=class_id).first()
    
    if not grading_weights:
        # Create default weights (25% each)
        grading_weights = GradingWeights(
            class_id=class_id,
            attendance_weight=25.0,
            instructor_participation_weight=25.0,
            peer_participation_weight=25.0,
            poll_weight=25.0
        )
        db.session.add(grading_weights)
        db.session.commit()
    
    return jsonify({
        'attendance_weight': grading_weights.attendance_weight,
        'instructor_participation_weight': grading_weights.instructor_participation_weight,
        'peer_participation_weight': grading_weights.peer_participation_weight,
        'poll_weight': grading_weights.poll_weight
    })

@app.route('/api/grading_weights/<int:class_id>', methods=['POST'])
@login_required
def update_grading_weights(class_id):
    class_obj = Class.query.get_or_404(class_id)
    if class_obj.professor_id != current_user.id:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    data = request.get_json()
    
    attendance_weight = float(data.get('attendance_weight', 25.0))
    instructor_participation_weight = float(data.get('instructor_participation_weight', 25.0))
    peer_participation_weight = float(data.get('peer_participation_weight', 25.0))
    poll_weight = float(data.get('poll_weight', 25.0))
    
    # Validate that weights sum to 100%
    total_weight = attendance_weight + instructor_participation_weight + peer_participation_weight + poll_weight
    if abs(total_weight - 100.0) > 0.01:  # Allow small floating point errors
        return jsonify({'success': False, 'error': f'Weights must sum to 100%. Current total: {total_weight}%'}), 400
    
    # Validate that all weights are non-negative
    if any(w < 0 for w in [attendance_weight, instructor_participation_weight, peer_participation_weight, poll_weight]):
        return jsonify({'success': False, 'error': 'All weights must be non-negative'}), 400
    
    grading_weights = GradingWeights.query.filter_by(class_id=class_id).first()
    
    if not grading_weights:
        grading_weights = GradingWeights(class_id=class_id)
        db.session.add(grading_weights)
    
    grading_weights.attendance_weight = attendance_weight
    grading_weights.instructor_participation_weight = instructor_participation_weight
    grading_weights.peer_participation_weight = peer_participation_weight
    grading_weights.poll_weight = poll_weight
    grading_weights.updated_at = datetime.utcnow()
    
    db.session.commit()
    
    # Recalculate all overall grades for this class with new weights (Task 14)
    # Note: Overall grades are calculated on-the-fly in get_gradebook, so no manual update needed
    # But we can trigger a socket event to notify clients to refresh if needed
    
    return jsonify({'success': True})

@app.route('/api/create_poll/<int:class_id>', methods=['POST'])
@login_required
def create_poll(class_id):
    class_obj = Class.query.get_or_404(class_id)
    if class_obj.professor_id != current_user.id:
        return jsonify({'success': False, 'error': 'Unauthorized'})
    
    data = request.get_json()
    question = data.get('question')
    options = data.get('options', [])
    correct_answer = data.get('correct_answer')
    is_graded = data.get('is_graded', False)
    is_anonymous = data.get('is_anonymous', False)
    
    # Validate options
    if not options or len(options) < 2:
        return jsonify({'success': False, 'error': 'At least 2 options are required'}), 400
    
    # Check for duplicate options (case-insensitive, trimmed)
    options_lower = [opt.strip().lower() for opt in options if opt]
    if len(options_lower) != len(set(options_lower)):
        return jsonify({'success': False, 'error': 'Duplicate options are not allowed. Each option must be unique.'}), 400
    
    # Deactivate any existing active polls
    Poll.query.filter_by(class_id=class_id, is_active=True).update({'is_active': False})
    
    poll = Poll(
        class_id=class_id,
        question=question,
        options=json.dumps(options),
        correct_answer=correct_answer,
        is_graded=is_graded,
        is_anonymous=is_anonymous,
        is_active=True
    )
    db.session.add(poll)
    db.session.commit()
    
    socketio.emit('poll_started', {
        'poll_id': poll.id,
        'question': question,
        'options': options,
        'is_graded': is_graded,
        'is_anonymous': is_anonymous
    }, room=f'class_{class_id}')
    
    return jsonify({'success': True, 'poll_id': poll.id})

@app.route('/api/stop_poll/<int:poll_id>', methods=['POST'])
@login_required
def stop_poll(poll_id):
    poll = Poll.query.get_or_404(poll_id)
    class_obj = Class.query.get_or_404(poll.class_id)
    if class_obj.professor_id != current_user.id:
        return jsonify({'success': False, 'error': 'Unauthorized'})
    
    poll.is_active = False
    db.session.commit()
    
    socketio.emit('poll_stopped', {'poll_id': poll_id}, room=f'class_{poll.class_id}')
    
    return jsonify({'success': True})

@app.route('/api/toggle_poll_graded/<int:poll_id>', methods=['POST'])
@login_required
def toggle_poll_graded(poll_id):
    poll = Poll.query.get_or_404(poll_id)
    class_obj = Class.query.get_or_404(poll.class_id)
    if class_obj.professor_id != current_user.id:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    data = request.get_json()
    is_graded = data.get('is_graded', not poll.is_graded)
    
    poll.is_graded = is_graded
    db.session.commit()
    
    return jsonify({'success': True, 'is_graded': poll.is_graded})

@app.route('/api/create_class', methods=['POST'])
@login_required
def create_class():
    data = request.get_json()
    name = data.get('name')
    class_code = data.get('class_code')
    
    if not name or not class_code:
        return jsonify({'success': False, 'error': 'Name and class code required'})
    
    if Class.query.filter_by(class_code=class_code).first():
        return jsonify({'success': False, 'error': 'Class code already exists'})
    
    class_obj = Class(
        professor_id=current_user.id,
        name=name,
        class_code=class_code
    )
    db.session.add(class_obj)
    db.session.commit()
    
    return jsonify({'success': True, 'class_id': class_obj.id})

@app.route('/api/delete_class/<int:class_id>', methods=['DELETE'])
@login_required
def delete_class(class_id):
    class_obj = Class.query.get_or_404(class_id)
    if class_obj.professor_id != current_user.id:
        return jsonify({'success': False, 'error': 'Unauthorized'})
    
    try:
        # Delete all related records
        # First, delete poll responses for polls in this class
        polls = Poll.query.filter_by(class_id=class_id).all()
        for poll in polls:
            PollResponse.query.filter_by(poll_id=poll.id).delete()
        
        # Delete polls
        Poll.query.filter_by(class_id=class_id).delete()
        
        # Delete participations
        Participation.query.filter_by(class_id=class_id).delete()
        
        # Delete attendances
        Attendance.query.filter_by(class_id=class_id).delete()
        
        # Delete enrollments
        Enrollment.query.filter_by(class_id=class_id).delete()
        
        # Delete class settings
        ClassSettings.query.filter_by(class_id=class_id).delete()
        
        # Finally, delete the class
        db.session.delete(class_obj)
        db.session.commit()
        
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/add_student_to_class', methods=['POST'])
@login_required
def add_student_to_class():
    data = request.get_json()
    class_id = data.get('class_id')
    student_id = data.get('student_id')
    
    class_obj = Class.query.get_or_404(class_id)
    if class_obj.professor_id != current_user.id:
        return jsonify({'success': False, 'error': 'Unauthorized'})
    
    enrollment = Enrollment.query.filter_by(
        class_id=class_id,
        student_id=student_id
    ).first()
    
    if enrollment:
        return jsonify({'success': False, 'error': 'Student already enrolled'})
    
    enrollment = Enrollment(class_id=class_id, student_id=student_id, is_active=True)
    db.session.add(enrollment)
    db.session.commit()
    
    return jsonify({'success': True})

@app.route('/api/create_and_add_student/<int:class_id>', methods=['POST'])
@login_required
def create_and_add_student(class_id):
    """Create a new student and add them to the class"""
    class_obj = Class.query.get_or_404(class_id)
    if class_obj.professor_id != current_user.id:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    data = request.get_json()
    
    # Safely extract and strip fields, handling None values
    first_name = (data.get('first_name') or '').strip() if data.get('first_name') is not None else ''
    last_name = (data.get('last_name') or '').strip() if data.get('last_name') is not None else ''
    preferred_name_raw = data.get('preferred_name')
    preferred_name = preferred_name_raw.strip() if preferred_name_raw else None
    student_number = (data.get('student_number') or '').strip() if data.get('student_number') is not None else ''
    email = (data.get('email') or '').strip() if data.get('email') is not None else ''
    rfid_card_id_raw = data.get('rfid_card_id')
    rfid_card_id = rfid_card_id_raw.strip() if rfid_card_id_raw else None
    
    # Validate required fields
    if not first_name or not last_name or not student_number or not email:
        return jsonify({'success': False, 'error': 'First name, last name, student number, and email are required'}), 400
    
    # Validate student number format (exactly 9 digits, no more, no less)
    if not student_number.isdigit():
        return jsonify({'success': False, 'error': 'Student number must contain only digits'}), 400
    
    if len(student_number) != 9:
        return jsonify({'success': False, 'error': f'Student number must be exactly 9 digits. You entered {len(student_number)} digit(s).'}), 400
    
    # Validate email format
    if '@' not in email:
        return jsonify({'success': False, 'error': 'Invalid email format'}), 400
    
    # Check if student already exists by student_number OR email (to ensure one account per student)
    existing_student = Student.query.filter_by(student_number=student_number).first()
    if not existing_student:
        # Also check by email to ensure uniqueness across classes
        existing_student = Student.query.filter_by(email=email).first()
    
    if existing_student:
        # Student exists, just enroll them if not already enrolled
        enrollment = Enrollment.query.filter_by(
            class_id=class_id,
            student_id=existing_student.id
        ).first()
        
        if enrollment:
            return jsonify({'success': False, 'error': 'Student is already enrolled in this class'}), 400
        
        # Update student info if provided (but don't change student_number or email if they differ)
        if existing_student.first_name != first_name:
            existing_student.first_name = first_name
        if existing_student.last_name != last_name:
            existing_student.last_name = last_name
        if existing_student.preferred_name != preferred_name:
            existing_student.preferred_name = preferred_name
        # Only update email if it matches (to maintain account uniqueness)
        if existing_student.email != email and existing_student.student_number == student_number:
            existing_student.email = email
        if rfid_card_id and existing_student.rfid_card_id != rfid_card_id:
            # Check if RFID is already taken by another student
            rfid_student = Student.query.filter_by(rfid_card_id=rfid_card_id).first()
            if rfid_student and rfid_student.id != existing_student.id:
                return jsonify({'success': False, 'error': 'RFID card ID already in use by another student'}), 400
            existing_student.rfid_card_id = rfid_card_id
        
        student = existing_student
    else:
        # Create new student - check if student_number or email conflicts
        # Check if student number is already taken
        if Student.query.filter_by(student_number=student_number).first():
            return jsonify({'success': False, 'error': 'Student number already exists'}), 400
        # Check if email is already taken
        if Student.query.filter_by(email=email).first():
            return jsonify({'success': False, 'error': 'Email already exists'}), 400
        # Check if RFID is already taken
        if rfid_card_id and Student.query.filter_by(rfid_card_id=rfid_card_id).first():
            return jsonify({'success': False, 'error': 'RFID card ID already in use'}), 400
        
        student = Student(
            student_number=student_number,
            first_name=first_name,
            last_name=last_name,
            preferred_name=preferred_name,
            email=email,
            rfid_card_id=rfid_card_id
        )
        db.session.add(student)
        db.session.flush()
    
    # Enroll student in class
    enrollment = Enrollment(class_id=class_id, student_id=student.id, is_active=True)
    db.session.add(enrollment)
    db.session.commit()
    
    return jsonify({
        'success': True,
        'student': {
            'id': student.id,
            'student_number': student.student_number,
            'first_name': student.first_name,
            'last_name': student.last_name,
            'preferred_name': student.preferred_name,
            'email': student.email
        }
    })

@app.route('/api/toggle_student_status/<int:class_id>/<int:student_id>', methods=['POST'])
@login_required
def toggle_student_status(class_id, student_id):
    class_obj = Class.query.get_or_404(class_id)
    if class_obj.professor_id != current_user.id:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    enrollment = Enrollment.query.filter_by(
        class_id=class_id,
        student_id=student_id
    ).first()
    
    if not enrollment:
        return jsonify({'success': False, 'error': 'Student not enrolled in this class'}), 404
    
    # Toggle the status
    enrollment.is_active = not enrollment.is_active
    db.session.commit()
    
    return jsonify({'success': True, 'is_active': enrollment.is_active})

@app.route('/api/update_student/<int:student_id>', methods=['POST'])
@login_required
def update_student(student_id):
    student = Student.query.get_or_404(student_id)
    
    # Check if any class the student is enrolled in belongs to current user
    enrollments = Enrollment.query.filter_by(student_id=student_id).all()
    if not enrollments:
        return jsonify({'success': False, 'error': 'Student not found in any of your classes'}), 404
    
    # Check if student is enrolled in at least one class owned by current user
    has_access = any(e.class_obj.professor_id == current_user.id for e in enrollments)
    if not has_access:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    data = request.get_json()
    
    first_name = data.get('first_name', '').strip()
    last_name = data.get('last_name', '').strip()
    preferred_name = data.get('preferred_name', '').strip() or None
    student_number = data.get('student_number', '').strip()
    email = data.get('email', '').strip()
    
    # Validate required fields
    if not first_name or not last_name or not student_number or not email:
        return jsonify({'success': False, 'error': 'First name, last name, student number, and email are required'}), 400
    
    # Validate student number format (9 digits)
    if len(student_number) != 9 or not student_number.isdigit():
        return jsonify({'success': False, 'error': 'Student number must be exactly 9 digits'}), 400
    
    # Validate email format
    if '@' not in email:
        return jsonify({'success': False, 'error': 'Invalid email format'}), 400
    
    # Check if student number is already taken by another student
    existing_student = Student.query.filter_by(student_number=student_number).first()
    if existing_student and existing_student.id != student_id:
        return jsonify({'success': False, 'error': 'Student number already exists'}), 400
    
    # Update student information
    student.first_name = first_name
    student.last_name = last_name
    student.preferred_name = preferred_name
    student.student_number = student_number
    student.email = email
    
    db.session.commit()
    
    return jsonify({'success': True})

@app.route('/api/remove_student_from_class/<int:class_id>/<int:student_id>', methods=['DELETE'])
@login_required
def remove_student_from_class(class_id, student_id):
    class_obj = Class.query.get_or_404(class_id)
    if class_obj.professor_id != current_user.id:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    student = Student.query.get_or_404(student_id)
    
    # Find enrollment
    enrollment = Enrollment.query.filter_by(
        class_id=class_id,
        student_id=student_id
    ).first()
    
    if not enrollment:
        return jsonify({'success': False, 'error': 'Student not enrolled in this class'}), 404
    
    # Delete all class-related data for this student in this class
    # Delete attendance records
    Attendance.query.filter_by(
        class_id=class_id,
        student_id=student_id
    ).delete()
    
    # Delete participation records
    Participation.query.filter_by(
        class_id=class_id,
        student_id=student_id
    ).delete()
    
    # Delete poll responses for polls in this class
    polls = Poll.query.filter_by(class_id=class_id).all()
    poll_ids = [poll.id for poll in polls]
    if poll_ids:
        PollResponse.query.filter(
            PollResponse.student_id == student_id,
            PollResponse.poll_id.in_(poll_ids)
        ).delete()
    
    # Delete enrollment
    db.session.delete(enrollment)
    
    db.session.commit()
    
    return jsonify({'success': True})

@app.route('/api/download_student_template/<int:class_id>')
@login_required
def download_student_template(class_id):
    class_obj = Class.query.get_or_404(class_id)
    if class_obj.professor_id != current_user.id:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    # Create a new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Students"
    
    # Define the header row with exact format (5 columns)
    # Column order: A=First Name, B=Last Name, C=Preferred Name, D=Student Number, E=Email
    headers = ['Student First Name', 'Last Name', 'Student Preferred Name', 'Student Number', 'Email']
    ws.append(headers)
    
    # Style the header row
    header_fill = PatternFill(start_color="2A1A40", end_color="2A1A40", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    # Add one example row with 9-digit student number (Student Number in column D, Email in column E)
    # Column order: First Name, Last Name, Preferred Name, Student Number, Email
    example_row = ['John', 'Doe', 'Johnny', '123456789', 'john.doe@example.com']
    ws.append(example_row)
    
    # Auto-fit column widths based on content
    from openpyxl.utils import get_column_letter
    
    # Calculate column widths based on header and example content
    # Excel column width units are approximately equal to the width of one character
    # We add extra padding for better readability
    # Columns: A=First Name, B=Last Name, C=Preferred Name, D=Student Number, E=Email
    columns_data = [
        ('A', 'Student First Name', example_row[0]),
        ('B', 'Last Name', example_row[1]),
        ('C', 'Student Preferred Name', example_row[2]),
        ('D', 'Student Number', example_row[3]),
        ('E', 'Email', example_row[4])
    ]
    
    for col_letter, header_text, example_text in columns_data:
        # Calculate width based on the longest content (header or example)
        # Multiply by 1.2 for padding and convert to Excel width units
        max_content_length = max(len(header_text), len(str(example_text)))
        # Excel width calculation: add padding and ensure minimum readable width
        column_width = max(max_content_length * 1.15 + 2, 12)
        # Cap maximum width at 50 to prevent extremely wide columns
        column_width = min(column_width, 50)
        ws.column_dimensions[col_letter].width = column_width
    
    # Create BytesIO object to store the Excel file in memory
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return send_file(
        excel_buffer,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'student_template_{class_obj.name.replace(" ", "_")}.xlsx'
    )

@app.route('/api/export_students/<int:class_id>')
@login_required
def export_students(class_id):
    class_obj = Class.query.get_or_404(class_id)
    if class_obj.professor_id != current_user.id:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    # Get active students only (matching template format)
    active_students = db.session.query(Student).join(Enrollment).filter(
        Enrollment.class_id == class_id,
        Enrollment.is_active == True
    ).order_by(Student.last_name, Student.first_name).all()
    
    # Create a new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Students"  # Match template sheet name
    
    # Define the header row with exact format (5 columns) - matching template exactly
    # Column order: A=First Name, B=Last Name, C=Preferred Name, D=Student Number, E=Email
    headers = ['Student First Name', 'Last Name', 'Student Preferred Name', 'Student Number', 'Email']
    ws.append(headers)
    
    # Style the header row - matching template styling
    header_fill = PatternFill(start_color="2A1A40", end_color="2A1A40", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    # Add active student rows with all 5 columns
    for student in active_students:
        row = [
            student.first_name,
            student.last_name,
            student.preferred_name if student.preferred_name else '',
            student.student_number,
            student.email if hasattr(student, 'email') and student.email else ''
        ]
        ws.append(row)
    
    # Auto-fit column widths - matching template calculation method
    from openpyxl.utils import get_column_letter
    
    if active_students:
        # Calculate max lengths including headers
        max_first_name = max([len(student.first_name) for student in active_students] + [len('Student First Name')])
        max_last_name = max([len(student.last_name) for student in active_students] + [len('Last Name')])
        max_preferred_name = max([len(student.preferred_name or '') for student in active_students] + [len('Student Preferred Name')])
        max_student_number = max([len(str(student.student_number)) for student in active_students] + [len('Student Number')])
        max_email = max([len(getattr(student, 'email', '') or '') for student in active_students] + [len('Email')])
    else:
        # If no students, use header lengths
        max_first_name = len('Student First Name')
        max_last_name = len('Last Name')
        max_preferred_name = len('Student Preferred Name')
        max_student_number = len('Student Number')
        max_email = len('Email')
    
    # Columns: A=First Name, B=Last Name, C=Preferred Name, D=Student Number, E=Email
    columns_data = [
        ('A', 'Student First Name', max_first_name),
        ('B', 'Last Name', max_last_name),
        ('C', 'Student Preferred Name', max_preferred_name),
        ('D', 'Student Number', max_student_number),
        ('E', 'Email', max_email)
    ]
    
    for col_letter, header_text, max_content_length in columns_data:
        # Calculate width based on the longest content (header or data)
        # Multiply by 1.15 for padding and add 2 for extra spacing
        column_width = max(max_content_length * 1.15 + 2, 12)
        # Cap maximum width at 50 to prevent extremely wide columns
        column_width = min(column_width, 50)
        ws.column_dimensions[col_letter].width = column_width
    
    # Create BytesIO object to store the Excel file in memory
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return send_file(
        excel_buffer,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'class_list_{class_obj.name.replace(" ", "_")}_{datetime.now().strftime("%Y%m%d")}.xlsx'
    )

@app.route('/api/upload_students/<int:class_id>', methods=['POST'])
@login_required
def upload_students(class_id):
    class_obj = Class.query.get_or_404(class_id)
    if class_obj.professor_id != current_user.id:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'No file uploaded'}), 400
    
    file = request.files['file']
    
    if file.filename == '':
        return jsonify({'success': False, 'error': 'No file selected'}), 400
    
    # Check if file is Excel
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'success': False, 'error': 'Invalid file type. Please upload an Excel file (.xlsx or .xls)'}), 400
    
    try:
        # Load the workbook
        wb = load_workbook(file, read_only=True, data_only=True)
        
        # Expected headers (5 columns) - Column order: First Name, Last Name, Preferred Name, Student Number, Email
        expected_headers = ['Student First Name', 'Last Name', 'Student Preferred Name', 'Student Number', 'Email']
        
        # Try to find "Active Students" sheet first, then fall back to active sheet
        ws_active = None
        ws_inactive = None
        
        if 'Active Students' in wb.sheetnames:
            ws_active = wb['Active Students']
        else:
            # Fall back to active sheet (for backward compatibility with single-sheet uploads)
            ws_active = wb.active
        
        # If "Inactive Students" sheet exists, use it (optional)
        if 'Inactive Students' in wb.sheetnames:
            ws_inactive = wb['Inactive Students']
        
        # Validate active sheet (required)
        if ws_active:
            header_row = []
            for cell in ws_active[1]:
                if cell.value is None:
                    header_row.append('')
                else:
                    header_row.append(str(cell.value).strip())
            
            header_row = header_row[:5]
            expected_normalized = [h.strip() for h in expected_headers]
            
            if header_row != expected_normalized:
                return jsonify({
                    'success': False, 
                    'error': f'Invalid file format. The "Active Students" sheet first row must match exactly: {expected_headers}. Your file has: {header_row}. Please download the template and use it exactly as provided (including exact column names and order).'
                }), 400
        
        # Validate inactive sheet if it exists (optional)
        if ws_inactive:
            header_row = []
            for cell in ws_inactive[1]:
                if cell.value is None:
                    header_row.append('')
                else:
                    header_row.append(str(cell.value).strip())
            
            header_row = header_row[:5]
            expected_normalized = [h.strip() for h in expected_headers]
            
            if header_row != expected_normalized:
                return jsonify({
                    'success': False, 
                    'error': f'Invalid file format. The "Inactive Students" sheet first row must match exactly: {expected_headers}. Your file has: {header_row}. Please download the template and use it exactly as provided (including exact column names and order).'
                }), 400
        
        # Process data rows (skip header row)
        results = {
            'success': True,
            'added': 0,
            'updated': 0,
            'activated': 0,
            'deactivated': 0,
            'skipped': 0,
            'errors': []
        }
        
        # Track all student numbers processed in this upload to handle duplicates within the upload
        processed_student_numbers_in_upload = set()
        
        # Get all existing student numbers in the database to avoid duplicate checks
        existing_student_numbers_db = {s.student_number for s in Student.query.all()}
        
        # Process Active Students sheet
        if ws_active:
            for row_idx, row in enumerate(ws_active.iter_rows(min_row=2, values_only=False), start=2):
                # Rule 1: Check that data only appears in columns A, B, C, D, E
                has_extra_data = False
                if len(row) > 5:
                    for cell in row[5:]:
                        if cell.value is not None and str(cell.value).strip():
                            has_extra_data = True
                            break
                
                if has_extra_data:
                    results['skipped'] += 1
                    results['errors'].append(f'Active Sheet Row {row_idx}: Data found beyond columns A-E. All data must be in columns A, B, C, D, and E only.')
                    continue
                
                # Extract values from first 5 columns (A, B, C, D, E)
                cell_values = []
                for cell in row[:5]:
                    if cell.value is None:
                        cell_values.append(None)
                    else:
                        raw_value = cell.value
                        cell_values.append(str(raw_value).strip() if raw_value else None)
                
                # Skip if all cells in first 5 columns are empty
                if not any(cell_values):
                    continue
                
                # Extract values based on column order:
                # A=First Name, B=Last Name, C=Preferred Name, D=Student Number, E=Email
                first_name = cell_values[0] if cell_values[0] else None
                last_name = cell_values[1] if cell_values[1] else None
                preferred_name_raw = cell_values[2] if cell_values[2] else None
                student_number_raw = cell_values[3] if cell_values[3] else None
                email_raw = cell_values[4] if cell_values[4] else None
                
                # Rule 5: Extract only digits from student number (ignore formatting)
                if student_number_raw:
                    student_number = ''.join(re.findall(r'\d', student_number_raw))
                else:
                    student_number = None
                
                # Preferred name: extract text
                if preferred_name_raw and preferred_name_raw.strip():
                    preferred_name = preferred_name_raw.strip()
                else:
                    preferred_name = None
                
                # Validate required fields
                if not student_number or not first_name or not last_name or not email_raw:
                    results['skipped'] += 1
                    results['errors'].append(f'Active Sheet Row {row_idx}: Missing required fields (Student Number, Student First Name, Last Name, or Email)')
                    continue
                
                # Validate email format
                email = email_raw.strip() if email_raw else None
                if not email or '@' not in email:
                    results['skipped'] += 1
                    results['errors'].append(f'Active Sheet Row {row_idx}: Invalid email format: "{email_raw}"')
                    continue
                
                # Rule 2: Validate Student Number is exactly 9 digits
                if len(student_number) != 9:
                    results['skipped'] += 1
                    results['errors'].append(f'Active Sheet Row {row_idx}: Student Number must be exactly 9 digits. Found {len(student_number)} digit(s) in: "{student_number_raw}"')
                    continue
                
                # Check for duplicates within this upload (skip if already processed in this upload)
                if student_number in processed_student_numbers_in_upload:
                    results['skipped'] += 1
                    results['errors'].append(f'Active Sheet Row {row_idx}: Duplicate student number "{student_number}" found in this upload. Only the first occurrence will be processed.')
                    continue
                
                # Mark this student number as processed in this upload
                processed_student_numbers_in_upload.add(student_number)
                
                # Normalize names
                first_name = first_name.strip() if first_name else None
                last_name = last_name.strip() if last_name else None
                
                # Check if student exists by student_number OR email (to ensure one account per student)
                existing_student = Student.query.filter_by(student_number=student_number).first()
                if not existing_student:
                    # Also check by email to ensure uniqueness across classes
                    existing_student = Student.query.filter_by(email=email).first()
                
                if existing_student:
                    # Update existing student if needed (but maintain account uniqueness)
                    updated = False
                    if existing_student.first_name != first_name:
                        existing_student.first_name = first_name
                        updated = True
                    if existing_student.last_name != last_name:
                        existing_student.last_name = last_name
                        updated = True
                    if existing_student.preferred_name != preferred_name:
                        existing_student.preferred_name = preferred_name
                        updated = True
                    # Only update email if student_number matches (to maintain account uniqueness)
                    if existing_student.email != email and existing_student.student_number == student_number:
                        existing_student.email = email
                        updated = True
                    
                    if updated:
                        results['updated'] += 1
                else:
                    # Create new student
                    new_student = Student(
                        student_number=student_number,
                        first_name=first_name,
                        preferred_name=preferred_name,
                        last_name=last_name,
                        email=email,
                        rfid_card_id=None
                    )
                    db.session.add(new_student)
                    db.session.flush()
                    existing_student = new_student
                    results['added'] += 1
                
                # Check if student is already enrolled in this class
                enrollment = Enrollment.query.filter_by(
                    class_id=class_id,
                    student_id=existing_student.id
                ).first()
                
                if not enrollment:
                    # Only create enrollment if student is not already enrolled
                    enrollment = Enrollment(class_id=class_id, student_id=existing_student.id, is_active=True)
                    db.session.add(enrollment)
                else:
                    # If re-enrolling an inactive student, activate them
                    if not enrollment.is_active:
                        enrollment.is_active = True
                        results['activated'] += 1
                    # If already enrolled and active, no action needed (not a duplicate error, just skip)
        
        # Process Inactive Students sheet (optional)
        if ws_inactive:
            for row_idx, row in enumerate(ws_inactive.iter_rows(min_row=2, values_only=False), start=2):
                # Rule 1: Check that data only appears in columns A, B, C, D, E
                has_extra_data = False
                if len(row) > 5:
                    for cell in row[5:]:
                        if cell.value is not None and str(cell.value).strip():
                            has_extra_data = True
                            break
                
                if has_extra_data:
                    results['skipped'] += 1
                    results['errors'].append(f'Inactive Sheet Row {row_idx}: Data found beyond columns A-E. All data must be in columns A, B, C, D, and E only.')
                    continue
                
                # Extract values from first 5 columns (A, B, C, D, E)
                cell_values = []
                for cell in row[:5]:
                    if cell.value is None:
                        cell_values.append(None)
                    else:
                        raw_value = cell.value
                        cell_values.append(str(raw_value).strip() if raw_value else None)
                
                # Skip if all cells in first 5 columns are empty
                if not any(cell_values):
                    continue
                
                # Extract values based on column order
                # A=First Name, B=Last Name, C=Preferred Name, D=Student Number, E=Email
                first_name = cell_values[0] if cell_values[0] else None
                last_name = cell_values[1] if cell_values[1] else None
                preferred_name_raw = cell_values[2] if cell_values[2] else None
                student_number_raw = cell_values[3] if cell_values[3] else None
                email_raw = cell_values[4] if cell_values[4] else None
                
                # Rule 5: Extract only digits from student number
                if student_number_raw:
                    student_number = ''.join(re.findall(r'\d', student_number_raw))
                else:
                    student_number = None
                
                # Skip if already processed in active sheet (active takes precedence)
                if student_number and student_number in processed_student_numbers_in_upload:
                    continue  # Skip - already processed in active sheet
                
                # Check for duplicates within inactive sheet (skip if already processed in this upload's inactive section)
                if student_number and student_number in processed_student_numbers_in_upload:
                    results['skipped'] += 1
                    results['errors'].append(f'Inactive Sheet Row {row_idx}: Duplicate student number "{student_number}" found in this upload. Only the first occurrence will be processed.')
                    continue
                
                # Preferred name: extract text
                if preferred_name_raw and preferred_name_raw.strip():
                    preferred_name = preferred_name_raw.strip()
                else:
                    preferred_name = None
                
                # Validate required fields
                if not student_number or not first_name or not last_name or not email_raw:
                    results['skipped'] += 1
                    results['errors'].append(f'Inactive Sheet Row {row_idx}: Missing required fields (Student Number, Student First Name, Last Name, or Email)')
                    continue
                
                # Validate email format
                email = email_raw.strip() if email_raw else None
                if not email or '@' not in email:
                    results['skipped'] += 1
                    results['errors'].append(f'Inactive Sheet Row {row_idx}: Invalid email format: "{email_raw}"')
                    continue
                
                # Rule 2: Validate Student Number is exactly 9 digits
                if len(student_number) != 9:
                    results['skipped'] += 1
                    results['errors'].append(f'Inactive Sheet Row {row_idx}: Student Number must be exactly 9 digits. Found {len(student_number)} digit(s) in: "{student_number_raw}"')
                    continue
                
                # Mark this student number as processed in this upload (for inactive sheet)
                processed_student_numbers_in_upload.add(student_number)
                
                # Normalize names
                first_name = first_name.strip() if first_name else None
                last_name = last_name.strip() if last_name else None
                
                # Check if student exists by student_number OR email (to ensure one account per student)
                existing_student = Student.query.filter_by(student_number=student_number).first()
                if not existing_student:
                    # Also check by email to ensure uniqueness across classes
                    existing_student = Student.query.filter_by(email=email).first()
                
                if existing_student:
                    # Update existing student if needed (but maintain account uniqueness)
                    updated = False
                    if existing_student.first_name != first_name:
                        existing_student.first_name = first_name
                        updated = True
                    if existing_student.last_name != last_name:
                        existing_student.last_name = last_name
                        updated = True
                    if existing_student.preferred_name != preferred_name:
                        existing_student.preferred_name = preferred_name
                        updated = True
                    # Only update email if student_number matches (to maintain account uniqueness)
                    if existing_student.email != email and existing_student.student_number == student_number:
                        existing_student.email = email
                        updated = True
                    
                    if updated:
                        results['updated'] += 1
                else:
                    # Create new student
                    new_student = Student(
                        student_number=student_number,
                        first_name=first_name,
                        preferred_name=preferred_name,
                        last_name=last_name,
                        email=email,
                        rfid_card_id=None
                    )
                    db.session.add(new_student)
                    db.session.flush()
                    existing_student = new_student
                    results['added'] += 1
                
                # Check if student is already enrolled in this class
                enrollment = Enrollment.query.filter_by(
                    class_id=class_id,
                    student_id=existing_student.id
                ).first()
                
                if not enrollment:
                    # Only create enrollment if student is not already enrolled
                    enrollment = Enrollment(class_id=class_id, student_id=existing_student.id, is_active=False)
                    db.session.add(enrollment)
                else:
                    # Mark as inactive if currently active
                    if enrollment.is_active:
                        enrollment.is_active = False
                        results['deactivated'] += 1
                    # If already enrolled and inactive, no action needed (not a duplicate error, just skip)
        
        # Commit all changes
        db.session.commit()
        
        return jsonify(results)
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': f'Error processing file: {str(e)}'}), 500

# Student routes
@app.route('/student')
def student_interface():
    html = render_template('student_interface.html')
    resp = make_response(html)
    # Prevent proxies/browsers from serving an old student UI after deploy
    resp.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    resp.headers['Pragma'] = 'no-cache'
    resp.headers['Expires'] = '0'
    return resp

def check_student_enrollment(student_id):
    """Check if student is enrolled in at least one class"""
    enrollment = Enrollment.query.filter_by(student_id=student_id).first()
    return enrollment is not None


def student_on_active_roster(student_id):
    """Student must appear on at least one class roster with active enrollment."""
    return Enrollment.query.filter_by(student_id=student_id, is_active=True).first() is not None


def _student_login_payload(stu):
    return {
        'id': stu.id,
        'student_number': stu.student_number,
        'first_name': stu.first_name,
        'last_name': stu.last_name,
        'preferred_name': stu.preferred_name,
        'email': stu.email,
    }


def _find_student_by_identifier(identifier):
    """Resolve student by 9-digit student number or email."""
    if not identifier or not str(identifier).strip():
        return None
    s = str(identifier).strip()
    if s.isdigit() and len(s) == 9:
        return Student.query.filter_by(student_number=s).first()
    return Student.query.filter_by(email=s).first()


@app.route('/api/student/login', methods=['POST'])
def student_login():
    data = request.get_json() or {}
    rfid_card_id = data.get('rfid_card_id')
    identifier = (data.get('identifier') or data.get('email') or '').strip()
    password = data.get('password') or ''

    student = None

    # RFID (hardware nameplates)
    if rfid_card_id:
        student = Student.query.filter_by(rfid_card_id=rfid_card_id.strip()).first()
        if student:
            if not student_on_active_roster(student.id):
                return jsonify({'success': False, 'error': 'You are not registered in any class. Please contact your professor.'})
            if not student.password_hash:
                session['student_id'] = student.id
                session['needs_password'] = True
                return jsonify({
                    'success': True,
                    'needs_password': True,
                    'student': _student_login_payload(student),
                })
            session['student_id'] = student.id
            session['needs_password'] = False
            return jsonify({
                'success': True,
                'needs_password': False,
                'student': {k: v for k, v in _student_login_payload(student).items() if k != 'email'},
            })

    # Student number or email + password
    if identifier:
        student = _find_student_by_identifier(identifier)
        if not student:
            return jsonify({'success': False, 'error': 'Invalid student number, email, or password.'})
        if not student_on_active_roster(student.id):
            return jsonify({'success': False, 'error': 'You are not registered in any class. Please contact your professor.'})
        if not student.password_hash:
            session['student_id'] = student.id
            session['needs_password'] = True
            return jsonify({
                'success': True,
                'needs_password': True,
                'student': _student_login_payload(student),
            })
        if not password:
            return jsonify({'success': False, 'error': 'Please enter your password.'})
        if check_password_hash(student.password_hash, password):
            session['student_id'] = student.id
            session['needs_password'] = False
            return jsonify({
                'success': True,
                'needs_password': False,
                'student': {k: v for k, v in _student_login_payload(student).items() if k != 'email'},
            })
        return jsonify({'success': False, 'error': 'Invalid student number, email, or password.'})

    return jsonify({'success': False, 'error': 'Please enter your student number or email and password.'})

@app.route('/api/student/find_for_password', methods=['POST'])
def find_student_for_password():
    """Find student by student number or email for password setup"""
    data = request.get_json()
    identifier = data.get('identifier', '').strip()
    
    if not identifier:
        return jsonify({'success': False, 'error': 'Please enter your student number or email'})
    
    # Try to find by student number (9 digits) or email
    student = None
    if identifier.isdigit() and len(identifier) == 9:
        student = Student.query.filter_by(student_number=identifier).first()
    else:
        student = Student.query.filter_by(email=identifier).first()
    
    if not student:
        return jsonify({'success': False, 'error': 'Student not found. Please make sure you are registered in a class by your professor.'})
    
    if not student_on_active_roster(student.id):
        return jsonify({'success': False, 'error': 'You are not registered in any class. Please contact your professor.'})
    
    # Set session for password setup
    session['student_id'] = student.id
    
    return jsonify({
        'success': True,
        'student': {
            'id': student.id,
            'student_number': student.student_number,
            'first_name': student.first_name,
            'last_name': student.last_name,
            'email': student.email,
            'has_password': student.password_hash is not None
        }
    })

@app.route('/api/student/set_password', methods=['POST'])
def student_set_password():
    """Set password for student on first login"""
    student_id = session.get('student_id')
    if not student_id:
        return jsonify({'success': False, 'error': 'Not authenticated. Please find your account first using the "Set Password / Register" option.'})
    
    data = request.get_json()
    password = data.get('password')
    confirm_password = data.get('confirm_password')
    
    if not password or not confirm_password:
        return jsonify({'success': False, 'error': 'Password and confirmation are required'})
    
    if password != confirm_password:
        return jsonify({'success': False, 'error': 'Passwords do not match'})
    
    if len(password) < 6:
        return jsonify({'success': False, 'error': 'Password must be at least 6 characters long'})
    
    student = Student.query.get(student_id)
    if not student:
        return jsonify({'success': False, 'error': 'Student not found'})
    
    if not student_on_active_roster(student.id):
        return jsonify({'success': False, 'error': 'You are not registered in any class. Please contact your professor.'})
    
    # Set the password
    student.password_hash = generate_password_hash(password)
    db.session.commit()
    
    session['needs_password'] = False
    
    return jsonify({
        'success': True,
        'message': 'Password set successfully',
        'student': {
            'id': student.id,
            'student_number': student.student_number,
            'first_name': student.first_name,
            'last_name': student.last_name,
            'preferred_name': student.preferred_name,
        }
    })

@app.route('/api/student/current', methods=['GET'])
def get_current_student():
    """Get current logged-in student info"""
    student_id = session.get('student_id')
    if not student_id:
        return jsonify({'success': False, 'error': 'Not logged in'}), 401
    
    student = Student.query.get(student_id)
    if not student:
        return jsonify({'success': False, 'error': 'Student not found'}), 404
    
    return jsonify({
        'success': True,
        'student': {
            'id': student.id,
            'student_number': student.student_number,
            'first_name': student.first_name,
            'last_name': student.last_name,
            'preferred_name': student.preferred_name,
            'email': student.email
        }
    })

@app.route('/api/student/classes')
def get_active_classes():
    """Active classes where student is on roster (enrollment active). Requires login."""
    student_id = session.get('student_id')
    if not student_id:
        return jsonify({'success': False, 'error': 'Not logged in'}), 401

    rows = db.session.query(Class).join(Enrollment).filter(
        Enrollment.student_id == student_id,
        Enrollment.is_active == True,
        Class.is_active == True,
    ).all()

    out = []
    for c in rows:
        settings = ClassSettings.query.filter_by(class_id=c.id).first()
        out.append({
            'id': c.id,
            'name': c.name,
            'class_code': c.class_code,
            'is_active': c.is_active,
            'show_first_name_only': bool(settings and settings.show_first_name_only),
        })
    return jsonify(out)

@app.route('/api/student/join_class', methods=['POST'])
def student_join_class():
    student_id = session.get('student_id')
    if not student_id:
        return jsonify({'success': False, 'error': 'Not logged in'})
    
    data = request.get_json()
    class_id = data.get('class_id')
    
    class_obj = Class.query.get_or_404(class_id)
    if not class_obj.is_active:
        return jsonify({'success': False, 'error': 'Class is not active'})

    enrollment = Enrollment.query.filter_by(
        class_id=class_id,
        student_id=student_id,
        is_active=True,
    ).first()
    if not enrollment:
        return jsonify({'success': False, 'error': 'You are not enrolled in this class.'})
    
    # Mark attendance and log join time
    today = datetime.utcnow().date()
    join_time = datetime.utcnow()
    attendance = Attendance.query.filter_by(
        class_id=class_id,
        student_id=student_id,
        date=today
    ).first()
    
    if not attendance:
        attendance = Attendance(
            class_id=class_id,
            student_id=student_id,
            date=today,
            present=True,
            join_time=join_time
        )
        db.session.add(attendance)
    else:
        # Update join_time if not set (in case of re-join)
        if not attendance.join_time:
            attendance.join_time = join_time
        attendance.present = True
    
    db.session.commit()
    
    socketio.emit('student_joined', {
        'student_id': student_id,
        'class_id': class_id
    }, room=f'class_{class_id}')
    
    return jsonify({'success': True, 'class_id': class_id})


@app.route('/api/student/leave_class', methods=['POST'])
def student_leave_class():
    """Record leave time for today without ending student session (return to class list)."""
    student_id = session.get('student_id')
    if not student_id:
        return jsonify({'success': False, 'error': 'Not logged in'})
    data = request.get_json() or {}
    class_id = data.get('class_id')
    if not class_id:
        return jsonify({'success': False, 'error': 'Class ID required'})
    today = datetime.utcnow().date()
    leave_time = datetime.utcnow()
    attendance = Attendance.query.filter_by(
        class_id=class_id,
        student_id=student_id,
        date=today,
    ).first()
    if attendance:
        attendance.leave_time = leave_time
        db.session.commit()
    return jsonify({'success': True})


@app.route('/api/student/logout', methods=['POST'])
def student_logout():
    student_id = session.get('student_id')
    if not student_id:
        return jsonify({'success': False, 'error': 'Not logged in'})
    
    data = request.get_json()
    class_id = data.get('class_id')
    
    if class_id:
        # Log leave time for this class
        today = datetime.utcnow().date()
        leave_time = datetime.utcnow()
        
        attendance = Attendance.query.filter_by(
            class_id=class_id,
            student_id=student_id,
            date=today
        ).first()
        
        if attendance:
            attendance.leave_time = leave_time
            db.session.commit()
    
    # Clear session
    session.pop('student_id', None)
    
    return jsonify({'success': True})

@app.route('/api/student/interaction', methods=['POST'])
def student_interaction():
    try:
        student_id = session.get('student_id')
        if not student_id:
            return jsonify({'success': False, 'error': 'Not logged in'})
        
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'error': 'No data provided'})
        
        class_id = data.get('class_id')
        interaction_type = data.get('type')  # 'hand_raise', 'thumbs_up', 'thumbs_down'
        
        if not class_id:
            return jsonify({'success': False, 'error': 'Class ID required'})
        
        if not interaction_type:
            return jsonify({'success': False, 'error': 'Interaction type required'})
        
        # Check if quiet mode is enabled
        settings = ClassSettings.query.filter_by(class_id=class_id).first()
        if settings and settings.quiet_mode and interaction_type in ['hand_raise', 'thumbs_up', 'thumbs_down']:
            return jsonify({'success': False, 'error': 'Quiet mode is enabled. Participation is disabled.'})
        
        today = datetime.utcnow().date()
        participation = Participation.query.filter_by(
            class_id=class_id,
            student_id=student_id,
            date=today
        ).first()
        
        if not participation:
            participation = Participation(
                class_id=class_id,
                student_id=student_id,
                date=today
            )
            db.session.add(participation)
        
        # Ensure fields are initialized to 0 if None
        if participation.hand_raises is None:
            participation.hand_raises = 0
        if participation.thumbs_up is None:
            participation.thumbs_up = 0
        if participation.thumbs_down is None:
            participation.thumbs_down = 0
        
        # Check current state for toggle behavior
        action = 'toggle'  # Default to toggle
        if 'action' in data:
            action = data.get('action')  # 'raise' or 'lower'
        
        if interaction_type == 'hand_raise':
            # Check if student already has an active (not cleared) hand raise
            active_hand_raise = HandRaise.query.filter_by(
                class_id=class_id,
                student_id=student_id,
                cleared=False
            ).first()
            
            if active_hand_raise:
                # Lower hand (toggle off)
                active_hand_raise.cleared = True
                # Don't increment participation count when lowering
            else:
                # Raise hand (toggle on)
                participation.hand_raises += 1
                hand_raise = HandRaise(
                    class_id=class_id,
                    student_id=student_id,
                    timestamp=datetime.utcnow()
                )
                db.session.add(hand_raise)
                
        elif interaction_type == 'thumbs_up':
            # Mutually exclusive with thumbs_down: only one can be "on" per student
            cu = participation.thumbs_up or 0
            cd = participation.thumbs_down or 0
            if cu % 2 == 1 and cd % 2 == 1:
                participation.thumbs_down = max(0, cd - 1)
                cd = participation.thumbs_down
            if cu % 2 == 1:
                participation.thumbs_up = max(0, cu - 1)
            else:
                if cd % 2 == 1:
                    participation.thumbs_down = max(0, cd - 1)
                participation.thumbs_up = cu + 1
        elif interaction_type == 'thumbs_down':
            cu = participation.thumbs_up or 0
            cd = participation.thumbs_down or 0
            if cu % 2 == 1 and cd % 2 == 1:
                participation.thumbs_up = max(0, cu - 1)
                cu = participation.thumbs_up
            if cd % 2 == 1:
                participation.thumbs_down = max(0, cd - 1)
            else:
                if cu % 2 == 1:
                    participation.thumbs_up = max(0, cu - 1)
                participation.thumbs_down = cd + 1
        else:
            return jsonify({'success': False, 'error': 'Invalid interaction type'})
        
        db.session.commit()
        
        # Determine if interaction is now active
        is_active = False
        if interaction_type == 'hand_raise':
            is_active = HandRaise.query.filter_by(
                class_id=class_id,
                student_id=student_id,
                cleared=False
            ).first() is not None
        elif interaction_type == 'thumbs_up':
            is_active = (participation.thumbs_up or 0) % 2 == 1
        elif interaction_type == 'thumbs_down':
            is_active = (participation.thumbs_down or 0) % 2 == 1
        
        # Refresh participation to get updated values
        db.session.refresh(participation)
        
        socketio.emit('student_interaction', {
            'student_id': student_id,
            'class_id': class_id,
            'type': interaction_type,
            'is_active': is_active
        }, room=f'class_{class_id}')
        
        out = {'success': True, 'is_active': is_active}
        if interaction_type in ('thumbs_up', 'thumbs_down'):
            out['thumbs_up_active'] = (participation.thumbs_up or 0) % 2 == 1
            out['thumbs_down_active'] = (participation.thumbs_down or 0) % 2 == 1
        return jsonify(out)
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/student/poll_response', methods=['POST'])
def student_poll_response():
    student_id = session.get('student_id')
    if not student_id:
        return jsonify({'success': False, 'error': 'Not logged in'})
    
    data = request.get_json()
    poll_id = data.get('poll_id')
    answer = data.get('answer')
    
    poll = Poll.query.get_or_404(poll_id)
    if not poll.is_active:
        return jsonify({'success': False, 'error': 'Poll is not active'})
    
    # Check if already responded
    existing = PollResponse.query.filter_by(
        poll_id=poll_id,
        student_id=student_id
    ).first()
    
    if existing:
        return jsonify({'success': False, 'error': 'Already responded'})
    
    is_correct = (poll.correct_answer is not None and answer == poll.correct_answer)
    
    response = PollResponse(
        poll_id=poll_id,
        student_id=student_id,
        answer=answer,
        is_correct=is_correct
    )
    db.session.add(response)
    db.session.commit()
    
    has_key = poll.correct_answer is not None
    socketio.emit('poll_response', {
        'poll_id': poll_id,
        'student_id': student_id,
        'answer': answer,
        'is_correct': is_correct,
        'is_anonymous': poll.is_anonymous
    }, room=f'class_{poll.class_id}')
    
    return jsonify({
        'success': True,
        'is_correct': is_correct,
        'has_correct_answer': has_key,
        'correct_answer_index': poll.correct_answer,
        'selected_index': answer,
    })

# Live Dashboard APIs
@app.route('/api/live_dashboard/<int:class_id>')
@login_required
def live_dashboard(class_id):
    class_obj = Class.query.get_or_404(class_id)
    if class_obj.professor_id != current_user.id:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    today = datetime.utcnow().date()
    
    # Get settings
    settings = ClassSettings.query.filter_by(class_id=class_id).first()
    if not settings:
        settings = ClassSettings(class_id=class_id)
        db.session.add(settings)
        db.session.commit()
    
    # Get total enrolled students
    total_students = db.session.query(Student).join(Enrollment).filter(
        Enrollment.class_id == class_id,
        Enrollment.is_active == True
    ).count()
    
    # Get present students
    present_students = db.session.query(Student).join(Attendance).filter(
        Attendance.class_id == class_id,
        Attendance.date == today,
        Attendance.present == True
    ).all()
    
    # Get hands raised (not cleared, ordered by timestamp)
    hands_raised = db.session.query(HandRaise, Student).join(Student).filter(
        HandRaise.class_id == class_id,
        HandRaise.cleared == False
    ).order_by(HandRaise.timestamp.asc()).all()
    
    hands_raised_list = []
    for hand_raise, student in hands_raised:
        hands_raised_list.append({
            'student_id': student.id,
            'student_number': student.student_number,
            'first_name': student.first_name,
            'last_name': student.last_name,
            'preferred_name': student.preferred_name,
            'timestamp': hand_raise.timestamp.isoformat()
        })
    
    # Count of currently active (not cleared) hand raises
    active_hand_raises_count = len(hands_raised_list)
    
    # Get participation counts
    participations = Participation.query.filter_by(
        class_id=class_id,
        date=today
    ).all()
    
    unique_participants = len(set(p.student_id for p in participations if (p.hand_raises or 0) + (p.thumbs_up or 0) + (p.thumbs_down or 0) > 0))
    
    # Get thumbs up/down counts (only from current session)
    active_session = ClassSession.query.filter_by(
        class_id=class_id,
        end_time=None
    ).order_by(ClassSession.start_time.desc()).first()
    
    thumbs_up_count = 0
    thumbs_down_count = 0
    if active_session:
        # Get thumbs from current session only - count only active (odd numbers)
        session_participations = Participation.query.filter_by(
            class_id=class_id,
            date=today
        ).all()
        # Count only students who have thumbs up/down active (odd count)
        thumbs_up_count = sum(1 for p in session_participations if (p.thumbs_up or 0) % 2 == 1)
        thumbs_down_count = sum(1 for p in session_participations if (p.thumbs_down or 0) % 2 == 1)
    
    # Get active poll
    active_poll = Poll.query.filter_by(class_id=class_id, is_active=True).first()
    poll_data = None
    poll_results_data = None
    if active_poll:
        poll_data = {
            'poll_id': active_poll.id,
            'question': active_poll.question,
            'options': json.loads(active_poll.options),
            'is_anonymous': active_poll.is_anonymous,
            'is_graded': active_poll.is_graded
        }
        
        # Include poll results in the same response to avoid dual refresh
        responses = PollResponse.query.filter_by(poll_id=active_poll.id).all()
        options = json.loads(active_poll.options)
        
        option_counts = {}
        for i in range(len(options)):
            option_counts[i] = sum(1 for r in responses if r.answer == i)
        
        poll_results_data = {
            'success': True,
            'question': active_poll.question,
            'options': options,
            'option_counts': option_counts,
            'total_responses': len(responses)
        }
    
    return jsonify({
        'success': True,
        'active_hand_raises_count': active_hand_raises_count,
        'hands_raised': hands_raised_list,
        'unique_participants': unique_participants,
        'thumbs_up_count': thumbs_up_count,
        'thumbs_down_count': thumbs_down_count,
        'present_students': len(present_students),
        'total_students': total_students,
        'active_poll': poll_data,
        'poll_results': poll_results_data,
        'show_first_name_only': settings.show_first_name_only
    })

@app.route('/api/live_attendance/<int:class_id>')
@login_required
def live_attendance(class_id):
    class_obj = Class.query.get_or_404(class_id)
    if class_obj.professor_id != current_user.id:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    today = datetime.utcnow().date()
    
    # Get all enrolled students
    enrolled_students = db.session.query(Student).join(Enrollment).filter(
        Enrollment.class_id == class_id,
        Enrollment.is_active == True
    ).all()
    
    # Get present students
    present_student_ids = set(
        db.session.query(Attendance.student_id).filter_by(
            class_id=class_id,
            date=today,
            present=True
        ).all()
    )
    present_student_ids = {id[0] for id in present_student_ids}
    
    # Get participation data for all students
    participations = Participation.query.filter_by(
        class_id=class_id,
        date=today
    ).all()
    participation_map = {p.student_id: p for p in participations}
    
    # Get poll grades
    poll_responses = PollResponse.query.join(Poll).filter(
        Poll.class_id == class_id,
        Poll.is_graded == True
    ).all()
    
    poll_grades = {}
    for response in poll_responses:
        student_id = response.student_id
        if student_id not in poll_grades:
            poll_grades[student_id] = {'correct': 0, 'total': 0}
        poll_grades[student_id]['total'] += 1
        if response.is_correct:
            poll_grades[student_id]['correct'] += 1
    
    present_students = []
    absent_students = []
    
    for student in enrolled_students:
        participation = participation_map.get(student.id)
        poll_grade_data = poll_grades.get(student.id, {'correct': 0, 'total': 0})
        poll_grade = (poll_grade_data['correct'] / poll_grade_data['total'] * 100) if poll_grade_data['total'] > 0 else 0
        
        student_data = {
            'id': student.id,
            'student_number': student.student_number,
            'first_name': student.first_name,
            'last_name': student.last_name,
            'preferred_name': student.preferred_name,
            'hand_raises': participation.hand_raises if participation else 0,
            'thumbs_up': participation.thumbs_up if participation else 0,
            'thumbs_down': participation.thumbs_down if participation else 0,
            'poll_grade': round(poll_grade, 1),
            'participation_freq': (participation.hand_raises or 0) + (participation.thumbs_up or 0) + (participation.thumbs_down or 0) if participation else 0
        }
        
        if student.id in present_student_ids:
            present_students.append(student_data)
        else:
            absent_students.append(student_data)
    
    return jsonify({
        'success': True,
        'present_students': present_students,
        'absent_students': absent_students
    })

@app.route('/api/live_preferences/<int:class_id>')
@login_required
def live_preferences(class_id):
    class_obj = Class.query.get_or_404(class_id)
    if class_obj.professor_id != current_user.id:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    settings = ClassSettings.query.filter_by(class_id=class_id).first()
    if not settings:
        settings = ClassSettings(class_id=class_id)
        db.session.add(settings)
        db.session.commit()
    
    # Get current session's exclude_from_grading status
    active_session = ClassSession.query.filter_by(
        class_id=class_id,
        end_time=None
    ).order_by(ClassSession.start_time.desc()).first()
    
    exclude_from_grading = active_session.exclude_from_grading if active_session else False
    
    return jsonify({
        'success': True,
        'quiet_mode': settings.quiet_mode,
        'show_first_name_only': settings.show_first_name_only,
        'exclude_from_grading': exclude_from_grading
    })

@app.route('/api/clear_hands_raised/<int:class_id>', methods=['POST'])
@login_required
def clear_hands_raised(class_id):
    class_obj = Class.query.get_or_404(class_id)
    if class_obj.professor_id != current_user.id:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    HandRaise.query.filter_by(
        class_id=class_id,
        cleared=False
    ).update({'cleared': True})
    
    db.session.commit()
    
    socketio.emit('all_hands_cleared', {'class_id': class_id}, room=f'class_{class_id}')
    
    return jsonify({'success': True})

@app.route('/api/clear_participation_count/<int:class_id>', methods=['POST'])
@login_required
def clear_participation_count(class_id):
    class_obj = Class.query.get_or_404(class_id)
    if class_obj.professor_id != current_user.id:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    # This is a conceptual reset - in practice, we might want to track this differently
    # For now, we'll just return success as the count is calculated from participations
    return jsonify({'success': True, 'message': 'Participation count is calculated from current session data'})

@app.route('/api/reset_thumbs_up/<int:class_id>', methods=['POST'])
@login_required
def reset_thumbs_up(class_id):
    class_obj = Class.query.get_or_404(class_id)
    if class_obj.professor_id != current_user.id:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    # Reset thumbs up for current session
    today = datetime.utcnow().date()
    participations = Participation.query.filter_by(
        class_id=class_id,
        date=today
    ).all()
    
    for p in participations:
        if p.thumbs_up and p.thumbs_up > 0:
            p.thumbs_up = 0
    
    db.session.commit()
    
    return jsonify({'success': True})

@app.route('/api/reset_thumbs_down/<int:class_id>', methods=['POST'])
@login_required
def reset_thumbs_down(class_id):
    class_obj = Class.query.get_or_404(class_id)
    if class_obj.professor_id != current_user.id:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    # Reset thumbs down for current session
    today = datetime.utcnow().date()
    participations = Participation.query.filter_by(
        class_id=class_id,
        date=today
    ).all()
    
    for p in participations:
        if p.thumbs_down and p.thumbs_down > 0:
            p.thumbs_down = 0
    
    db.session.commit()
    
    return jsonify({'success': True})

@app.route('/api/student/interaction_state/<int:class_id>')
def get_student_interaction_state(class_id):
    """Get current interaction states for logged-in student"""
    student_id = session.get('student_id')
    if not student_id:
        return jsonify({'success': False, 'error': 'Not logged in'}), 401
    
    # Check hand raise
    hand_raise_active = HandRaise.query.filter_by(
        class_id=class_id,
        student_id=student_id,
        cleared=False
    ).first() is not None
    
    # Check thumbs up/down
    today = datetime.utcnow().date()
    participation = Participation.query.filter_by(
        class_id=class_id,
        student_id=student_id,
        date=today
    ).first()
    
    thumbs_up_active = False
    thumbs_down_active = False
    if participation:
        thumbs_up_active = (participation.thumbs_up or 0) % 2 == 1
        thumbs_down_active = (participation.thumbs_down or 0) % 2 == 1
    
    return jsonify({
        'success': True,
        'hand_raise_active': hand_raise_active,
        'thumbs_up_active': thumbs_up_active,
        'thumbs_down_active': thumbs_down_active
    })

@app.route('/api/poll_results/<int:poll_id>')
@login_required
def poll_results(poll_id):
    poll = Poll.query.get_or_404(poll_id)
    class_obj = Class.query.get_or_404(poll.class_id)
    if class_obj.professor_id != current_user.id:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    responses = PollResponse.query.filter_by(poll_id=poll_id).all()
    options = json.loads(poll.options)
    
    option_counts = {}
    for i in range(len(options)):
        option_counts[i] = sum(1 for r in responses if r.answer == i)
    
    return jsonify({
        'success': True,
        'question': poll.question,
        'options': options,
        'option_counts': option_counts,
        'total_responses': len(responses)
    })

# SocketIO Events
@socketio.on('connect')
def on_connect():
    emit('connected', {'data': 'Connected'})

@socketio.on('join_class')
def on_join_class(data):
    class_id = data.get('class_id')
    join_room(f'class_{class_id}')
    emit('joined_class', {'class_id': class_id})


@socketio.on('leave_class')
def on_leave_class(data):
    if data and data.get('class_id') is not None:
        leave_room(f'class_{data["class_id"]}')

@socketio.on('get_live_stats')
def on_get_live_stats(data):
    class_id = data.get('class_id')
    
    students = db.session.query(Student).join(Enrollment).filter(
        Enrollment.class_id == class_id,
        Enrollment.is_active == True
    ).all()
    
    today = datetime.utcnow().date()
    present_students = db.session.query(Student).join(Attendance).filter(
        Attendance.class_id == class_id,
        Attendance.date == today,
        Attendance.present == True
    ).all()
    
    participations = Participation.query.filter_by(
        class_id=class_id,
        date=today
    ).all()
    
    total_hand_raises = sum(p.hand_raises for p in participations)
    total_thumbs_up = sum(p.thumbs_up for p in participations)
    total_thumbs_down = sum(p.thumbs_down for p in participations)
    
    active_poll = Poll.query.filter_by(class_id=class_id, is_active=True).first()
    poll_stats = None
    if active_poll:
        responses = PollResponse.query.filter_by(poll_id=active_poll.id).all()
        option_counts = {}
        for i in range(len(json.loads(active_poll.options))):
            option_counts[i] = sum(1 for r in responses if r.answer == i)
        poll_stats = {
            'poll_id': active_poll.id,
            'question': active_poll.question,
            'options': json.loads(active_poll.options),
            'option_counts': option_counts,
            'total_responses': len(responses),
            'is_anonymous': active_poll.is_anonymous
        }
    
    emit('live_stats', {
        'total_students': len(students),
        'present_students': len(present_students),
        'total_hand_raises': total_hand_raises,
        'total_thumbs_up': total_thumbs_up,
        'total_thumbs_down': total_thumbs_down,
        'poll_stats': poll_stats
    })

def migrate_database():
    """Add missing columns and tables to existing database."""
    from sqlalchemy import inspect, text
    
    try:
        inspector = inspect(db.engine)
        table_names = inspector.get_table_names()
        
        # Check if student table exists and add missing columns
        if 'student' in table_names:
            student_columns = [col['name'] for col in inspector.get_columns('student')]
            
            if 'preferred_name' not in student_columns:
                try:
                    db.session.execute(text('ALTER TABLE student ADD COLUMN preferred_name VARCHAR(100)'))
                    db.session.commit()
                    print("✓ Added preferred_name column to student table")
                except Exception as e:
                    db.session.rollback()
                    print(f"✗ Error adding preferred_name column: {e}")
            
            if 'email' not in student_columns:
                try:
                    # SQLite doesn't support DEFAULT in ALTER TABLE, so add as nullable first
                    # We'll update existing records with a placeholder email
                    db.session.execute(text('ALTER TABLE student ADD COLUMN email VARCHAR(120)'))
                    db.session.commit()
                    
                    # Update existing records with a placeholder email if they don't have one
                    # Format: student_number@placeholder.local
                    students_without_email = db.session.execute(
                        text('SELECT id, student_number FROM student WHERE email IS NULL OR email = ""')
                    ).fetchall()
                    
                    for student_id, student_number in students_without_email:
                        placeholder_email = f"{student_number}@placeholder.local"
                        db.session.execute(
                            text('UPDATE student SET email = :email WHERE id = :id'),
                            {'email': placeholder_email, 'id': student_id}
                        )
                    
                    db.session.commit()
                    print("✓ Added email column to student table")
                    if students_without_email:
                        print(f"  → Updated {len(students_without_email)} existing students with placeholder emails")
                        print("  → Please update student emails via Excel import or manual edit")
                except Exception as e:
                    db.session.rollback()
                    print(f"✗ Error adding email column: {e}")
            
            if 'password_hash' not in student_columns:
                try:
                    db.session.execute(text('ALTER TABLE student ADD COLUMN password_hash VARCHAR(255)'))
                    db.session.commit()
                    print("✓ Added password_hash column to student table")
                except Exception as e:
                    db.session.rollback()
                    print(f"✗ Error adding password_hash column: {e}")
        
        # Check if attendance table exists and add missing columns
        if 'attendance' in table_names:
            attendance_columns = [col['name'] for col in inspector.get_columns('attendance')]
            
            if 'join_time' not in attendance_columns:
                try:
                    db.session.execute(text('ALTER TABLE attendance ADD COLUMN join_time DATETIME'))
                    db.session.commit()
                    print("✓ Added join_time column to attendance table")
                except Exception as e:
                    db.session.rollback()
                    print(f"✗ Error adding join_time column: {e}")
            
            if 'leave_time' not in attendance_columns:
                try:
                    db.session.execute(text('ALTER TABLE attendance ADD COLUMN leave_time DATETIME'))
                    db.session.commit()
                    print("✓ Added leave_time column to attendance table")
                except Exception as e:
                    db.session.rollback()
                    print(f"✗ Error adding leave_time column: {e}")
        
        # Check if poll table exists and add is_graded column
        if 'poll' in table_names:
            poll_columns = [col['name'] for col in inspector.get_columns('poll')]
            if 'is_graded' not in poll_columns:
                try:
                    db.session.execute(text('ALTER TABLE poll ADD COLUMN is_graded BOOLEAN DEFAULT 0'))
                    db.session.commit()
                    print("✓ Added is_graded column to poll table")
                except Exception as e:
                    db.session.rollback()
                    print(f"✗ Error adding is_graded column: {e}")
        
        # Check if class_session table exists and add exclude_from_grading column
        if 'class_session' in table_names:
            session_columns = [col['name'] for col in inspector.get_columns('class_session')]
            if 'exclude_from_grading' not in session_columns:
                try:
                    db.session.execute(text('ALTER TABLE class_session ADD COLUMN exclude_from_grading BOOLEAN DEFAULT 0'))
                    db.session.commit()
                    print("✓ Added exclude_from_grading column to class_session table")
                except Exception as e:
                    db.session.rollback()
                    print(f"✗ Error adding exclude_from_grading column: {e}")
        
        # Check if enrollment table exists and if is_active column exists
        if 'enrollment' in table_names:
            enrollment_columns = [col['name'] for col in inspector.get_columns('enrollment')]
            if 'is_active' not in enrollment_columns:
                try:
                    db.session.execute(text('ALTER TABLE enrollment ADD COLUMN is_active BOOLEAN DEFAULT 1'))
                    db.session.commit()
                    print("✓ Added is_active column to enrollment table")
                except Exception as e:
                    db.session.rollback()
                    print(f"✗ Error adding is_active column: {e}")
        
        # Check if hand_raise table exists
        if 'hand_raise' not in table_names:
            try:
                # Table will be created by create_all
                print("✓ hand_raise table will be created")
            except Exception as e:
                print(f"✗ Error with hand_raise table: {e}")
        
        # Ensure all tables exist - create_all will handle new tables
        db.create_all()
        print("✓ Database migration completed")
    except Exception as e:
        print(f"✗ Error during database migration: {e}")

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
        migrate_database()
        
        # Create a default professor for testing
        if not Professor.query.first():
            default_prof = Professor(
                username='professor',
                email='prof@example.com',
                password_hash=generate_password_hash('password')
            )
            db.session.add(default_prof)
            db.session.commit()
    
    socketio.run(app, debug=True, host='0.0.0.0', port=5000)

